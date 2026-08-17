import pandas as pd
import logging
from pathlib import Path
import duckdb 
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config.settings import DATA_DIR, EXCEL_FILE, DATABASE_PATH

def _verify_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

def load_to_duckdb(frame: pd.DataFrame) -> None:

    if frame.empty:
        logging.warning("Data Frame is empty, skipping duckdb")
        return
    
    _verify_dir()

    con = duckdb.connect(str(DATABASE_PATH))

    try:
        con.execute("""
            CREATE TABLE IF NOT EXISTS weather_observations (
                city VARCHAR,
                country VARCHAR,
                weather VARCHAR,
                description VARCHAR,
                temperature DOUBLE,
                feels_like DOUBLE,
                humidity BIGINT,
                windspeed DOUBLE,
                timestamp TIMESTAMP,
                extracted_at TIMESTAMP
            )
        """)

        # Rename columns to match DuckDB schema exactly
        db_frame = frame.rename(columns={
            "City": "city",
            "Country": "country",
            "Weather": "weather",
            "Description": "description",
            "Temperature": "temperature",
            "Feels_like": "feels_like",
            "Humidity": "humidity",
            "Wind_speed": "windspeed",
            "Timestamp": "timestamp",
            "Extracted_At": "extracted_at"
        })

        con.execute("INSERT INTO weather_observations SELECT * FROM db_frame")
        logging.info("Succesfully loaded %s records into DuckDB", len(frame))

    finally:
        con.close()
def load_to_excel(frame: pd.DataFrame) -> None:

    if frame.empty:
        logging.warning("Data Frame is empty, skipping excel report")
        return

    _verify_dir()

    if EXCEL_FILE.exists():
        existing = pd.read_excel(EXCEL_FILE)
        combined = pd.concat([existing, frame], ignore_index=True)
        combined.drop_duplicates(inplace=True, subset=['City','Timestamp'])
    else:
        combined = frame
    
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        combined.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            for cell in col:
                if cell.row == 1:
                    continue  # Skip header alignment
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif cell.value and ("timestamp" in str(worksheet.cell(1, cell.column).value).lower() or "extracted_at" in str(worksheet.cell(1, cell.column).value).lower()):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
    
    logging.info("Excel report successfully created")
    
def load_all(frame: pd.DataFrame) -> None:
    load_to_duckdb(frame)
    load_to_excel(frame)